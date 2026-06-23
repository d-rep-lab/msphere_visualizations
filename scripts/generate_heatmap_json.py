"""
Assumption: two-sheet Excel sheet!

Converts the two-sheet Excel workbook into carbon_heatmap_data.json.

Sheet1: time-series OD readings — row 0 is timepoints, column 0 is "<strain> <carbon>" labels.
Sheet2: lookup grid — row 0 is carbon sources, column A is strains.

Missing OD values default to 0. Strain/carbon order follows what is written in Sheet2.
"""

import json
import sys
import openpyxl

XLSX_PATH = "../data/total data.xlsx"
OUT_PATH = "../data/heatmap.json"


def load_lookup(wb):
    # Return (strains, carbons) lists from Sheet2, ordered as in the sheet.

    ws = wb["Sheet2"]
    rows = list(ws.iter_rows(values_only=True))
    carbons = [c for c in rows[0][1:] if c is not None]
    strains = [r[0] for r in rows[1:] if r[0] is not None]
    return strains, carbons


def build_parser(strains):
    # Return a function that splits '<strain> <carbon>' using the known list.

    ordered = sorted(
        strains, key=len, reverse=True
    )  # Sort longest-first so multi-word strains are before shorter prefixes.

    def parse(label):
        label = label.replace("\xa0", " ").strip()
        for strain in ordered:
            if label.startswith(strain + " "):
                return strain, label[len(strain) :].strip()
        raise ValueError(f"No known strain matches label: {label!r}")

    return parse


def parse_timeseries(ws, parse_label):
    # Return (timepoints, raw_data) from Sheet1.
    all_rows = list(ws.iter_rows(values_only=True))
    timepoints = [x for x in all_rows[0] if x is not None]
    n_time = len(timepoints)

    raw_data = []
    for row in all_rows[1:]:
        if row[0] is None:
            continue
        try:
            strain, carbon = parse_label(str(row[0]))
        except ValueError as exc:
            print(f"  Skipping row: {exc}", file=sys.stderr)
            continue

        values = [
            round(float(v), 3) if v is not None else 0.0 for v in row[1 : n_time + 1]
        ]
        values += [0.0] * (n_time - len(values))
        raw_data.append({"strain": strain, "carbon": carbon, "od": values})

    return timepoints, raw_data


def build_output(raw_data, timepoints, strains_order, carbons_order):
    # Assemble the JSON-ready output dict from parsed rows.

    all_peaks = [max(d["od"]) for d in raw_data]
    cells, series = [], []
    for d in raw_data:
        label = f"{d['strain']}|{d['carbon']}"
        cells.append(
            {
                "strain": d["strain"],
                "carb": d["carbon"],
                "max": round(max(d["od"]), 3),
                "series_label": label,
            }
        )
        series.append(
            {
                "label": label,
                "values": [{"t": t, "v": v} for t, v in zip(timepoints, d["od"])],
            }
        )

    return {
        "times": timepoints,
        "rows": strains_order,
        "cols": carbons_order,
        "global_min": round(min(all_peaks), 3),
        "global_max": round(max(all_peaks), 3),
        "cells": cells,
        "series": series,
    }


def write_output(output, path):
    """Write output dict to JSON and print a summary line."""

    with open(path, "w") as f:
        json.dump(output, f, indent=2)
    n_rows, n_cols = len(output["rows"]), len(output["cols"])
    print(
        f"{n_rows} strains x {n_cols} carbons = {len(output['cells'])} cells → {path}"
    )


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    strains_order, carbons_order = load_lookup(wb)
    parse_label = build_parser(strains_order)

    timepoints, raw_data = parse_timeseries(wb["Sheet1"], parse_label)
    if not raw_data:
        sys.exit("No data rows parsed. Check sheet names and label format.")

    output = build_output(raw_data, timepoints, strains_order, carbons_order)
    write_output(output, OUT_PATH)


if __name__ == "__main__":
    main()
