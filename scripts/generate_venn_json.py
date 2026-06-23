"""
Assumption: Uses the first sheet of the Excel file.

Sheet1: time-series OD readings — row 0 is timepoints, column 0 is "<strain> <substrate>" labels.
Strain names are parsed by matching known two-word prefixes ("Type 3") before single-word strains.

Missing OD values default to 0. Strain/substrate order follows the first strain encountered in Sheet1.
"""

import json
import sys
import openpyxl

XLSX_PATH = "../data/total data.xlsx"
OUT_PATH = "../data/venn.json"


def parse_row_name(name):
    # Return (strain, substrate) from a row label, or None if unparseable.
    if not name:
        return None
    # "Type 3" is the only two-word strain name.
    if name.startswith("Type 3 "):
        return "Type 3", name[7:]
    parts = name.split(" ", 1)
    if len(parts) == 2:
        return parts[0], parts[1]
    return None


def load_excel(path):
    # Return parsed data dict from Sheet1.
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    time_points = [float(h) for h in header_row[1:] if h is not None]
    n_times = len(time_points)

    data = {}
    strain_order = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        parsed = parse_row_name(row[0])
        if parsed is None:
            continue
        strain, substrate = parsed

        if strain not in data:
            data[strain] = {}
            strain_order.append(strain)

        values = [float(v) if v is not None else 0.0 for v in row[1 : n_times + 1]]
        data[strain][substrate] = values

    if not strain_order:
        sys.exit("No data rows parsed. Check sheet name and label format.")

    substrate_order = list(data[strain_order[0]].keys())

    return {
        "timePoints": time_points,
        "strains": strain_order,
        "substrates": substrate_order,
        "data": data,
    }


def write_output(output, path):
    """Write output dict to JSON and print a summary line."""

    with open(path, "w") as f:
        json.dump(output, f, indent=2)
    n_strains = len(output["strains"])
    n_substrates = len(output["substrates"])
    print(f"{n_strains} strains x {n_substrates} substrates → {path}")


def main():
    result = load_excel(XLSX_PATH)
    write_output(result, OUT_PATH)


if __name__ == "__main__":
    main()
